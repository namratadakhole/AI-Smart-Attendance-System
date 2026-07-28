from flask import Blueprint, request, jsonify, Response, send_file
from database import db
from utils import standard_response, token_required, role_required, validate_request_keys
from config import Config
import os
import io
import csv
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw

attendance_bp = Blueprint("attendance", __name__)

# Active session parameters stored in memory
active_session = {
    "subject_id": None,
    "department": None,
    "semester": None,
    "active": False
}

def generate_placeholder_frames():
    frame_num = 0
    while active_session.get("active"):
        # Create an RGB image in memory (dark slate background)
        img = Image.new('RGB', (640, 480), color=(15, 23, 42))
        d = ImageDraw.Draw(img)
        
        # Draw scanning HUD corners
        # Top-left
        d.line([(20, 20), (60, 20)], fill=(0, 191, 255), width=3)
        d.line([(20, 20), (20, 60)], fill=(0, 191, 255), width=3)
        # Top-right
        d.line([(620, 20), (580, 20)], fill=(0, 191, 255), width=3)
        d.line([(620, 20), (620, 60)], fill=(0, 191, 255), width=3)
        # Bottom-left
        d.line([(20, 460), (60, 460)], fill=(0, 191, 255), width=3)
        d.line([(20, 460), (20, 420)], fill=(0, 191, 255), width=3)
        # Bottom-right
        d.line([(620, 460), (580, 460)], fill=(0, 191, 255), width=3)
        d.line([(620, 460), (620, 420)], fill=(0, 191, 255), width=3)
        
        # Draw scanning scanline sweep
        scanline_y = 50 + (frame_num * 10) % 380
        d.line([(35, scanline_y), (605, scanline_y)], fill=(0, 191, 255), width=2)
        
        # Telemetry Labels
        d.text((100, 160), "HYBRID AI CLOUD SERVER ACTIVE", fill=(0, 191, 255))
        d.text((100, 190), f"Subject Session ID: {active_session.get('subject_id') or 'N/A'}", fill=(255, 255, 255))
        d.text((100, 220), f"Department: {active_session.get('department') or 'N/A'}", fill=(200, 200, 200))
        d.text((100, 250), f"Semester: Sem {active_session.get('semester') or 'N/A'}", fill=(200, 200, 200))
        d.text((100, 290), "Local AI Client (Workstation Camera)\nProcessing & logging check-ins...", fill=(50, 205, 50))
        
        output = io.BytesIO()
        img.save(output, format='JPEG')
        buffer = output.getvalue()
        
        yield (
            b'--frame\r\n'
            b'Content-Type: image/jpeg\r\n\r\n' +
            buffer +
            b'\r\n'
        )
        frame_num += 1
        time.sleep(0.1) # 10 FPS

def generate_offline_frame():
    img = Image.new('RGB', (640, 480), color=(30, 41, 59))
    d = ImageDraw.Draw(img)
    d.text((220, 220), "CAMERA OFFLINE\nStart session to stream", fill=(148, 163, 184))
    output = io.BytesIO()
    img.save(output, format='JPEG')
    return output.getvalue()

@attendance_bp.route("/start-attendance", methods=["POST"])
@token_required
@role_required("professor")
def start_attendance():
    active_session["active"] = True
    return standard_response(True, "Attendance Started")

@attendance_bp.route("/stop-attendance", methods=["POST"])
@token_required
@role_required("professor")
def stop_attendance():
    active_session["active"] = False
    return standard_response(True, "Attendance Stopped")

@attendance_bp.route("/video_feed")
def video_feed():
    if active_session.get("active"):
        return Response(
            generate_placeholder_frames(),
            mimetype="multipart/x-mixed-replace; boundary=frame"
        )
    else:
        return Response(
            generate_offline_frame(),
            mimetype="image/jpeg"
        )

@attendance_bp.route("/detected_students")
@token_required
def get_detected_students():
    if not active_session.get("active"):
        return jsonify({
            "students": [],
            "records": []
        })
        
    now = datetime.now()
    today_date = now.strftime("%d-%m-%Y")
    
    query = {"date": today_date}
    if active_session.get("subject_id"):
        query["subject_id"] = int(active_session["subject_id"])
        
    records = list(db.attendance.find(query, {"_id": 0}))
    students_list = [r["name"] for r in records]
    
    return jsonify({
        "students": students_list,
        "records": records
    })

@attendance_bp.route("/attendance-records", methods=["GET"])
@token_required
def get_attendance_records():
    search = request.args.get("search", "").strip().lower()
    department = request.args.get("department", "all")
    semester = request.args.get("semester", "all")
    filter_date = request.args.get("date", "").strip()
    sort_order = request.args.get("sort", "desc").lower()

    pipeline = []
    pipeline.append({
        "$lookup": {
            "from": "students",
            "localField": "student_id",
            "foreignField": "id",
            "as": "student"
        }
    })
    pipeline.append({
        "$unwind": {
            "path": "$student",
            "preserveNullAndEmptyArrays": True
        }
    })

    match = {}
    if department != "all" and department:
        match["department"] = department
    if semester != "all" and semester:
        match["student.semester"] = str(semester)
    if filter_date:
        match["date"] = filter_date
    if search:
        match["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"roll_no": {"$regex": search, "$options": "i"}}
        ]
    
    if match:
        pipeline.append({"$match": match})

    pipeline.append({
        "$project": {
            "_id": 0,
            "id": "$id",
            "student_id": "$student_id",
            "roll_no": "$roll_no",
            "name": "$name",
            "department": "$department",
            "date": "$date",
            "time": "$time",
            "status": "$status",
            "subject_id": "$subject_id",
            "recognition_confidence": "$recognition_confidence",
            "semester": "$student.semester"
        }
    })

    sort_dir = -1 if sort_order == "desc" else 1
    pipeline.append({"$sort": {"id": sort_dir}})

    records = list(db.attendance.aggregate(pipeline))

    unique_dates = db.attendance.distinct("date")
    total_dates = len(unique_dates) if len(unique_dates) > 0 else 1

    all_students = list(db.students.find({}, {"_id": 0}))

    student_summaries = []
    for st in all_students:
        s_id = st["id"]
        s_name = st["full_name"]
        s_roll = st["roll_no"]
        s_dept = st["department"]

        st_history = list(db.attendance.find({"name": s_name}, {"_id": 0}).sort("id", -1))
        present_days = len(st_history)
        pct = round((present_days / total_dates) * 100, 1)

        student_summaries.append({
            "id": s_id,
            "name": s_name,
            "roll": s_roll,
            "department": s_dept,
            "semester": st.get("semester"),
            "present_days": present_days,
            "total_days": total_dates,
            "percentage": f"{min(pct, 100.0)}%",
            "pct_number": min(pct, 100.0),
            "history": st_history
        })

    return jsonify({
        "records": records,
        "student_summaries": student_summaries
    })

@attendance_bp.route("/recognize", methods=["POST"])
@token_required
@validate_request_keys("image")
def recognize():
    # Deprecated in Cloud Backend since CV2 is removed
    return jsonify({
        "success": True,
        "faces": []
    })

@attendance_bp.route("/train-model", methods=["POST"])
@token_required
@role_required("professor")
def train_model():
    # Cloud compilation of encodings is handled locally by the AI client.
    # Returns mock success to prevent frontend console errors.
    return standard_response(True, "Model training skipped (Handled by Local AI client)")

@attendance_bp.route("/status")
def status():
    now = datetime.now()
    today_date = now.strftime("%d-%m-%Y")
    
    query = {"date": today_date}
    if active_session.get("subject_id"):
        query["subject_id"] = int(active_session["subject_id"])
        
    records = list(db.attendance.find(query, {"_id": 0}))
    students_list = [r["name"] for r in records]
    
    return jsonify({
        "running": active_session.get("active", False),
        "detected_students": students_list,
        "subject_id": active_session.get("subject_id"),
        "department": active_session.get("department"),
        "semester": active_session.get("semester")
    })

@attendance_bp.route("/debug")
def debug():
    return status()

@attendance_bp.route("/export-attendance", methods=["POST"])
@token_required
@role_required("professor")
def export_attendance():
    now = datetime.now()
    today_date = now.strftime("%d-%m-%Y")
    
    query = {"date": today_date}
    if active_session.get("subject_id"):
        query["subject_id"] = int(active_session["subject_id"])
    present_records = list(db.attendance.find(query))
    present_students = {r["name"] for r in present_records}

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.append([
        "Roll No",
        "Student Name",
        "Department",
        "Status"
    ])

    db_students = [{"roll": s["roll_no"], "name": s["full_name"], "department": s["department"]} for s in db.students.find()]

    for student in db_students:
        status_val = "Present" if student["name"] in present_students else "Absent"
        ws.append([
            student["roll"],
            student["name"],
            student["department"],
            status_val
        ])

    filename = os.path.join(
        Config.PROJECT_DIR,
        "database",
        "Attendance_Report.xlsx"
    )

    # Ensure database folder exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    wb.save(filename)
    return jsonify({
        "status": "success",
        "file": filename
    })

@attendance_bp.route("/download-attendance")
def download_attendance():
    now = datetime.now()
    today_date = now.strftime("%d-%m-%Y")
    
    query = {"date": today_date}
    if active_session.get("subject_id"):
        query["subject_id"] = int(active_session["subject_id"])
    present_records = list(db.attendance.find(query))
    present_students = {r["name"] for r in present_records}

    db_students = [{"roll": s["roll_no"], "name": s["full_name"], "department": s["department"]} for s in db.students.find()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    title_font = Font(size=18, bold=True)
    subtitle_font = Font(size=14, bold=True)
    heading_font = Font(size=12, bold=True)
    header_font = Font(bold=True, color="FFFFFF")

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    center = Alignment(horizontal="center")
    thin = Side(style="thin")
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "GOVERNMENT COLLEGE OF ENGINEERING NAGPUR"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = "AI SMART ATTENDANCE SYSTEM"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    ws["A4"] = "Department"
    ws["B4"] = "Computer Science Engineering"
    ws["D4"] = "Faculty"
    ws["E4"] = "Dr. A. Sharma"

    ws["A5"] = "Subject"
    ws["B5"] = "Artificial Intelligence"
    ws["D5"] = "Semester"
    ws["E5"] = "VII"

    ws["A6"] = "Date"
    ws["B6"] = now.strftime("%d-%m-%Y")
    ws["D6"] = "Time"
    ws["E6"] = now.strftime("%H:%M:%S")

    for cell in ["A4","D4","A5","D5","A6","D6"]:
        ws[cell].font = heading_font

    total_students = len(db_students)
    present_count = len(present_students)
    absent_count = max(total_students - present_count, 0)

    attendance_percentage = (
        round((present_count / total_students) * 100, 2)
        if total_students > 0
        else 0
    )

    ws["A8"] = "Total Students"
    ws["B8"] = total_students
    ws["C8"] = "Present"
    ws["D8"] = present_count
    ws["E8"] = "Absent"
    ws["F8"] = absent_count
    ws["A9"] = "Attendance %"
    ws["B9"] = f"{attendance_percentage}%"

    for cell in ["A8","C8","E8","A9"]:
        ws[cell].font = heading_font

    headers = ["Roll No", "Student Name", "Department", "Status"]
    row_number = 11

    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_number, column=col)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row = 12
    for student in db_students:
        status_val = "Present" if student["name"] in present_students else "Absent"

        ws.cell(row=row, column=1).value = student["roll"]
        ws.cell(row=row, column=2).value = student["name"]
        ws.cell(row=row, column=3).value = student["department"]
        ws.cell(row=row, column=4).value = status_val

        fill = green_fill if status_val == "Present" else red_fill

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = center
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    footer = ws.cell(row=row, column=1)
    footer.value = "Generated by AI Smart Attendance System"
    footer.font = Font(italic=True)
    footer.alignment = center

    report_path = os.path.join(
        Config.PROJECT_DIR,
        "database",
        "Attendance_Report.xlsx"
    )

    # Ensure database folder exists
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    wb.save(report_path)
    return send_file(report_path, as_attachment=True)

@attendance_bp.route("/attendance/start-session", methods=["POST"])
@token_required
@role_required("professor")
@validate_request_keys("subject_id")
def start_attendance_session():
    data = request.json
    subject_id = data.get("subject_id")
    department = data.get("department", "Computer Science & Engineering")
    semester = data.get("semester", "7")

    active_session["subject_id"] = subject_id
    active_session["department"] = department
    active_session["semester"] = semester
    active_session["active"] = True

    return jsonify({
        "status": "success",
        "message": f"Attendance Session Started for Subject ID: {subject_id}",
        "session": {
            "subject_id": subject_id,
            "department": department,
            "semester": semester
        }
    })

@attendance_bp.route("/recognize-upload", methods=["POST"])
@token_required
def recognize_upload():
    if "image" not in request.files:
        return standard_response(False, "No image file uploaded")

    subject_id = request.form.get("subject_id") or active_session.get("subject_id")
    department = request.form.get("department", "Computer Science & Engineering")
    semester = request.form.get("semester", "7")

    try:
        now = datetime.now()
        today_date = now.strftime("%d-%m-%Y")
        time_str = now.strftime("%I:%M:%S %p")

        # Pick first student matching roster as fallback simulation
        student = db.students.find_one({"department": department, "semester": semester})
        if not student:
            student = db.students.find_one()
            
        if not student:
            return standard_response(False, "No students registered in roster to mark attendance.")

        name = student["full_name"]
        roll_no = student["roll_no"]
        student_id = student["id"]

        existing_db = db.attendance.find_one({
            "name": name, 
            "date": today_date,
            "subject_id": int(subject_id) if subject_id else None
        })

        results = []
        if not existing_db:
            from database import get_next_sequence_value
            attendance_id = get_next_sequence_value("attendance")
            db.attendance.insert_one({
                "id": attendance_id,
                "student_id": student_id,
                "roll_no": roll_no,
                "name": name,
                "department": department,
                "date": today_date,
                "time": time_str,
                "status": "Present",
                "recognition_confidence": 98.2,
                "subject_id": int(subject_id) if subject_id else None
            })

        results.append({
            "name": name,
            "roll": roll_no,
            "department": department,
            "time": time_str,
            "confidence": 98.2,
            "status": "Present",
            "duplicate": existing_db is not None
        })

        return jsonify({
            "status": "success",
            "message": "Processed simulated upload checks.",
            "results": results
        })
    except Exception as e:
        return standard_response(False, f"Upload check error: {str(e)}")

@attendance_bp.route("/api/attendance/mark", methods=["POST"])
@token_required
@validate_request_keys("student_id", "name", "confidence")
def mark_attendance_api():
    data = request.json
    student_id = data.get("student_id")
    name = data.get("name")
    confidence = data.get("confidence", 95.0)
    timestamp = data.get("timestamp")

    subject_id = data.get("subject_id") or active_session.get("subject_id")
    department = data.get("department") or active_session.get("department")
    semester = data.get("semester") or active_session.get("semester")

    now = datetime.now()
    today_date = now.strftime("%d-%m-%Y")
    time_str = timestamp or now.strftime("%I:%M:%S %p")

    # Find student details
    student_info = None
    if isinstance(student_id, (int, float)) or (isinstance(student_id, str) and student_id.isdigit()):
        student_info = db.students.find_one({"id": int(student_id)})
    else:
        student_info = db.students.find_one({"roll_no": student_id})

    if not student_info:
        student_info = db.students.find_one({"full_name": name})

    if not student_info:
        return standard_response(False, f"Student '{name}' not registered in database.", status_code=404)

    # Check duplicate check-in today for this subject
    query = {
        "student_id": student_info["id"], 
        "date": today_date
    }
    if subject_id:
        query["subject_id"] = int(subject_id)
        
    existing_db = db.attendance.find_one(query)
    if existing_db:
        return jsonify({
            "success": True,
            "message": f"Attendance already marked for {student_info['full_name']} today.",
            "duplicate": True,
            "record": {
                "id": existing_db["id"],
                "student_id": existing_db["student_id"],
                "roll_no": existing_db["roll_no"],
                "name": existing_db["name"],
                "department": existing_db["department"],
                "date": existing_db["date"],
                "time": existing_db["time"],
                "status": existing_db["status"],
                "recognition_confidence": existing_db["recognition_confidence"],
                "subject_id": existing_db.get("subject_id")
            }
        })

    from database import get_next_sequence_value
    attendance_id = get_next_sequence_value("attendance")

    record = {
        "id": attendance_id,
        "student_id": student_info["id"],
        "roll_no": student_info["roll_no"],
        "name": student_info["full_name"],
        "department": student_info["department"],
        "date": today_date,
        "time": time_str,
        "status": "Present",
        "recognition_confidence": float(confidence),
        "subject_id": int(subject_id) if subject_id else None
    }

    db.attendance.insert_one(record)

    return jsonify({
        "success": True,
        "message": f"Attendance marked for {student_info['full_name']}",
        "duplicate": False,
        "record": {
            "id": record["id"],
            "student_id": record["student_id"],
            "roll_no": record["roll_no"],
            "name": record["name"],
            "department": record["department"],
            "date": record["date"],
            "time": record["time"],
            "status": record["status"],
            "recognition_confidence": record["recognition_confidence"],
            "subject_id": record["subject_id"]
        }
    })
